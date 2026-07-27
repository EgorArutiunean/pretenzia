from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.modules.excel_normalizer.build_debt_registry_template import build_registry, load_object_addresses


def _append_row(worksheet, level: int, values: list[object]) -> None:
    worksheet.append(values)
    worksheet.row_dimensions[worksheet.max_row].outlineLevel = level


class ExcelNormalizerTests(unittest.TestCase):
    def test_build_registry_uses_object_address_dictionary_and_writes_errors(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            raw_path = temp / "raw.xlsx"
            addresses_path = temp / "object_addresses.json"
            output_path = temp / "registry.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Лист_1"
            worksheet.append(["Организация", None, None, None, "Начислено", "Оплачено", "Долг на конец периода"])
            _append_row(worksheet, 2, ["Тестовая организация", None, None, None, 3000, 0, 3000])
            _append_row(worksheet, 3, ["12970001 от 01.01.2026", None, None, None, 1000, 0, 1000])
            _append_row(worksheet, 4, ["Иванов Иван", None, None, None, 1000, 0, 1000])
            _append_row(worksheet, 5, ["Январь 2026", None, None, None, 1000, 0, 1000])
            _append_row(worksheet, 3, ["99990002 от 01.01.2026", None, None, None, 2000, 0, 2000])
            _append_row(worksheet, 4, ["Петров Петр", None, None, None, 2000, 0, 2000])
            _append_row(worksheet, 5, ["Январь 2026", None, None, None, 2000, 0, 2000])
            workbook.save(raw_path)

            addresses_path.write_text(json.dumps({"1297": "Москва, Тестовая улица, д. 1"}, ensure_ascii=False), encoding="utf-8")

            rows = build_registry(
                input_path=str(raw_path),
                output_path=str(output_path),
                object_addresses_path=str(addresses_path),
            )
            result = load_workbook(output_path, data_only=True)
            registry = result["Реестр"]
            errors = result["Ошибки"]

        self.assertEqual(len(rows), 1)
        self.assertEqual(registry.max_row - 1, 1)
        self.assertEqual(
            registry.cell(2, 3).value,
            "Москва, Тестовая улица, д. 1, машиноместо № 0001",
        )
        self.assertEqual(errors.max_row - 1, 1)
        self.assertIn("9999", errors.cell(2, 6).value)

    def test_load_object_addresses_rejects_directory(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "должен быть файлом"):
                load_object_addresses(temp_dir)

    def test_registry_uses_company_data_and_tracks_non_positive_debt(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            raw_path = temp / "raw.xlsx"
            base_path = temp / "base.xlsx"
            output_path = temp / "registry.xlsx"

            raw = Workbook()
            raw_ws = raw.active
            raw_ws.append(
                ["Организация", None, None, None, "Начислено", "Оплачено", "Долг на конец периода"]
            )
            _append_row(raw_ws, 3, ["12970001 от 01.01.2026", None, None, None, 1000, 0, 1000])
            _append_row(raw_ws, 4, ["Иванов Иван", None, None, None, 1000, 0, 1000])
            _append_row(raw_ws, 5, ["Январь 2026", None, None, None, 1000, 0, 1000])
            _append_row(raw_ws, 3, ["12970002 от 01.01.2026", None, None, None, 0, 0, 0])
            _append_row(raw_ws, 4, ["Петров Петр", None, None, None, 0, 0, 0])
            raw.save(raw_path)

            base = Workbook()
            base_ws = base.active
            base_ws.append(
                [
                    "Номер объекта",
                    "Адрес",
                    "Компания",
                    "ИНН компании",
                    "Код компании",
                ]
            )
            base_ws.append(
                ["1297", "Москва, Тестовая улица, д. 1", "ООО Тест", "7700000000", "TEST"]
            )
            base.save(base_path)

            rows = build_registry(
                input_path=str(raw_path),
                output_path=str(output_path),
                base_data_path=str(base_path),
            )
            result = load_workbook(output_path, data_only=True)
            registry = result["Реестр"]
            good_payers = result["Добросовестные плательщики"]

        self.assertEqual(len(rows), 1)
        self.assertEqual(registry.cell(2, 6).value, "ООО Тест")
        self.assertEqual(registry.cell(2, 7).value, "7700000000")
        self.assertEqual(registry.cell(2, 8).value, "TEST")
        self.assertEqual(good_payers.max_row - 1, 1)
        self.assertEqual(good_payers.cell(2, 2).value, "12970002")

    def test_invalid_debt_is_reported_instead_of_becoming_zero(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            raw_path = temp / "raw.xlsx"
            output_path = temp / "registry.xlsx"

            raw = Workbook()
            raw_ws = raw.active
            raw_ws.append(
                ["Организация", None, None, None, "Начислено", "Оплачено", "Долг на конец периода"]
            )
            _append_row(
                raw_ws,
                3,
                ["12970001 от 01.01.2026", None, None, None, 1000, 0, "ошибка"],
            )
            _append_row(raw_ws, 4, ["Иванов Иван", None, None, None, 1000, 0, "ошибка"])
            raw.save(raw_path)

            rows = build_registry(
                input_path=str(raw_path),
                output_path=str(output_path),
                object_address="Москва",
            )
            result = load_workbook(output_path, data_only=True)
            errors = result["Ошибки"]

        self.assertEqual(rows, [])
        self.assertEqual(errors.max_row - 1, 1)
        self.assertIn("Не удалось распознать", errors.cell(2, 6).value)


if __name__ == "__main__":
    unittest.main()
