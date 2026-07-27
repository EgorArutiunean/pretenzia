from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COLUMN_ALIASES = {
    "номер объекта": "object_code",
    "номер обьекта": "object_code",
    "код объекта": "object_code",
    "код обьекта": "object_code",
    "объект": "object_code",
    "обьект": "object_code",
    "адрес": "object_address",
    "адрес объекта": "object_address",
    "компания": "company",
    "организация": "company",
    "код компании": "company_code",
    "инн компании": "company_inn",
    "инн": "company_inn",
    "реквизиты": "requisites",
    "реквезиты": "requisites",
    "генеральный директор": "director_name",
    "директор": "director_name",
    "ставка": "monthly_rate",
    "протокол": "protocol",
}


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower()).replace("ё", "е")


def _digits(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _extract_regex(text: str, pattern: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
    return " ".join(match.group(1).split()) if match else ""


def _extract_between(text: str, start: str, end: str) -> str:
    return _extract_regex(
        text,
        re.escape(start) + r"\s+(.+?)\s+" + re.escape(end),
    )


def _parse_rate(value: Any, row_idx: int) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        rate = Decimal(str(value).replace("\u00a0", "").replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Основная БЗ, строка {row_idx}: некорректная ставка.") from exc
    if rate != rate.to_integral_value():
        raise ValueError(
            f"Основная БЗ, строка {row_idx}: ставка должна быть указана без копеек."
        )
    return rate


@dataclass(frozen=True)
class ObjectData:
    object_code: str
    object_address: str
    company: str
    company_code: str
    company_inn: str
    requisites: str
    director_name_value: str
    protocol: str
    monthly_rate: Decimal | None

    @property
    def inn(self) -> str:
        return self.company_inn or _extract_regex(self.requisites, r"ИНН\s+(\d+)")

    @property
    def kpp(self) -> str:
        return _extract_regex(self.requisites, r"КПП\s+(\d+)")

    @property
    def ogrn(self) -> str:
        return _extract_regex(self.requisites, r"ОГРН\s+(\d+)")

    @property
    def legal_address(self) -> str:
        return _extract_between(
            self.requisites,
            "Юридический адрес",
            "Генеральный директор",
        )

    @property
    def director_name(self) -> str:
        if self.director_name_value:
            return self.director_name_value
        value = _extract_between(
            self.requisites,
            "Генеральный директор",
            "Главный бухгалтер",
        )
        if value:
            return value
        return _extract_regex(
            self.requisites,
            r"Генеральный директор\s+(.+?)(?:\s+Расчетный счет|$)",
        )


def load_object_data(path: str | Path) -> dict[str, ObjectData]:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Основная БЗ не найдена: {source}")
    if not source.is_file():
        raise ValueError(f"Путь к основной БЗ должен указывать на Excel-файл: {source}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Основная БЗ должна иметь формат .xlsx или .xlsm.")

    try:
        worksheet = load_workbook(source, data_only=True).active
    except Exception as exc:
        raise ValueError("Не удалось прочитать основную БЗ как Excel-файл.") from exc

    columns: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        canonical = COLUMN_ALIASES.get(normalize_header(worksheet.cell(1, col_idx).value))
        if canonical:
            columns[canonical] = col_idx

    required = {
        "object_code": "Номер объекта",
        "object_address": "Адрес",
        "company": "Компания",
    }
    missing = [label for key, label in required.items() if key not in columns]
    if missing:
        raise ValueError(
            "В основной БЗ не найдены обязательные колонки: " + ", ".join(missing)
        )

    result: dict[str, ObjectData] = {}
    row_by_code: dict[str, int] = {}
    for row_idx in range(2, worksheet.max_row + 1):
        raw_code = worksheet.cell(row_idx, columns["object_code"]).value
        digits = _digits(raw_code)
        if not digits:
            continue
        if len(digits) > 4:
            raise ValueError(
                f"Основная БЗ, строка {row_idx}: код объекта должен состоять из 4 цифр."
            )
        object_code = digits.zfill(4)
        if object_code in result:
            raise ValueError(
                f"В основной БЗ повторяется код объекта {object_code}: "
                f"строки {row_by_code[object_code]} и {row_idx}."
            )

        def value(name: str) -> Any:
            col_idx = columns.get(name)
            return worksheet.cell(row_idx, col_idx).value if col_idx else None

        result[object_code] = ObjectData(
            object_code=object_code,
            object_address=str(value("object_address") or "").strip(),
            company=str(value("company") or "").strip(),
            company_code=str(value("company_code") or "").strip(),
            company_inn=_digits(value("company_inn")),
            requisites=str(value("requisites") or "").strip(),
            director_name_value=str(value("director_name") or "").strip(),
            protocol=str(value("protocol") or "").strip(),
            monthly_rate=_parse_rate(value("monthly_rate"), row_idx),
        )
        row_by_code[object_code] = row_idx

    if not result:
        raise ValueError("В основной БЗ нет строк с кодами объектов.")
    return result
