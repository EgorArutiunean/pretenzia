from __future__ import annotations

import asyncio
import logging
import uuid
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

from aiogram import Bot, F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, FSInputFile, Message
from openpyxl import load_workbook

from app.bot.keyboards import data_updates_menu_keyboard, main_menu_keyboard
from app.config import load_settings
from app.modules.court_orders.generate_court_orders_from_registry import (
    WorkbookValidationReport,
    validate_base_data,
    validate_jurisdiction_data,
)
from app.modules.excel_normalizer.build_debt_registry_template import load_object_addresses
from app.pipeline import (
    PROJECT_ROOT,
    run_excel_to_registry,
    run_registry_to_claims,
    run_registry_to_court_orders,
)
from app.shared.reference_files import install_validated_reference_file


router = Router()
logger = logging.getLogger(__name__)
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
DEFAULT_OBJECT_ADDRESSES_PATH = PROJECT_ROOT / "storage" / "object_addresses.xlsx"
DEFAULT_COURT_ORDERS_BASE_DATA_PATH = PROJECT_ROOT / "storage" / "court_orders" / "base_data.xlsx"
DEFAULT_COURT_ORDERS_JURISDICTION_PATH = PROJECT_ROOT / "storage" / "court_orders" / "jurisdiction.xlsx"


class DocumentFlow(StatesGroup):
    waiting_for_file = State()


ACTION_PROMPTS = {
    "normalize": "Загрузите Excel-файл ОНВ из 1С.",
    "claims": (
        "Загрузите готовый Excel-реестр с колонками: Лицевой счет, ФИО, "
        "Адрес, Период задолженности, Сумма долга."
    ),
    "dictionary": "Загрузите Excel-справочник адресов объектов.",
    "court_data": "Загрузите основную Excel-БЗ: объекты, компании, реквизиты, протоколы и ставки.",
    "jurisdiction": "Загрузите Excel-БЗ подсудности с кодами объектов и судебными участками.",
    "court_orders": (
        "Загрузите готовый Excel-реестр с колонками: Лицевой счет, ФИО, "
        "Адрес, Период задолженности, Сумма долга."
    ),
}


HELP_TEXT = (
    "КАК РАБОТАТЬ\n"
    "\n"
    "1. Обновление данных\n"
    "Раздел доступен администраторам. Откройте его, когда изменился справочник:\n"
    "• Справочник адресов — коды объектов и адреса;\n"
    "• Основная БЗ — компании, реквизиты, протоколы и ставки;\n"
    "• Подсудность — судебные участки и адреса судов.\n"
    "После загрузки бот покажет количество объектов и предупреждения.\n"
    "\n"
    "2. Создать реестр из ОНВ\n"
    "Отправьте исходный Excel-отчет из 1С. Бот вернет registry.xlsx. "
    "Проверьте лист «Ошибки»: строки с ошибками не попадут в документы.\n"
    "\n"
    "3. Создать претензии из реестра\n"
    "Отправьте registry.xlsx. Обязательные колонки: Лицевой счет, ФИО, "
    "Адрес, Период задолженности, Сумма долга. Бот вернет claims.zip.\n"
    "\n"
    "4. Создать заявления в суд\n"
    "Сначала должны быть загружены Основная БЗ и Подсудность. Затем отправьте "
    "registry.xlsx. Для полного заявления также нужны: Дата рождения, Место рождения, "
    "Адрес регистрации и Паспорт или Идентификатор должника. Заявление формируется "
    "только при рассчитанной сумме от 5 000 до 500 000 рублей. Пени пока не "
    "рассчитываются. Бот вернет court_orders.zip.\n"
    "\n"
    "Если в court_orders.zip есть errors.xlsx, обязательно проверьте листы "
    "«Ошибки» и «Предупреждения» перед использованием документов.\n"
    "\n"
    "Поддерживаются Excel-файлы .xlsx и .xlsm. Для возврата в начало отправьте /start."
)


REFERENCE_ACTIONS = {"dictionary", "court_data", "jurisdiction"}


def _is_admin(user_id: int | None) -> bool:
    return user_id is not None and user_id in load_settings().admin_ids


def _is_allowed(user_id: int | None) -> bool:
    if user_id is None:
        return False
    return user_id in load_settings().allowed_user_ids


def _main_menu(user_id: int | None):
    return main_menu_keyboard(is_admin=_is_admin(user_id))


async def _deny_if_needed(message: Message) -> bool:
    if _is_allowed(message.from_user.id if message.from_user else None):
        return False
    user_id = message.from_user.id if message.from_user else None
    suffix = f"\nВаш Telegram ID: {user_id}" if user_id is not None else ""
    await message.answer("Доступ запрещён." + suffix)
    return True


async def _deny_callback_if_needed(callback: CallbackQuery) -> bool:
    if _is_allowed(callback.from_user.id if callback.from_user else None):
        return False
    await _answer_callback_safely(callback, "Доступ запрещён.", show_alert=True)
    return True


async def _deny_admin_message_if_needed(message: Message) -> bool:
    user_id = message.from_user.id if message.from_user else None
    if _is_admin(user_id):
        return False
    await message.answer("Это действие доступно только администратору.")
    return True


async def _deny_admin_callback_if_needed(callback: CallbackQuery) -> bool:
    if _is_admin(callback.from_user.id if callback.from_user else None):
        return False
    await _answer_callback_safely(
        callback,
        "Это действие доступно только администратору.",
        show_alert=True,
    )
    return True


async def _answer_callback_safely(
    callback: CallbackQuery,
    text: str | None = None,
    *,
    show_alert: bool = False,
) -> None:
    try:
        await callback.answer(text=text, show_alert=show_alert)
    except TelegramBadRequest as exc:
        if "query is too old" not in str(exc).lower():
            raise
        logger.info("Skipping expired callback query id=%s", callback.id)


def _create_run_dir(user_id: int) -> Path:
    run_id = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_" + uuid.uuid4().hex[:8]
    run_dir = PROJECT_ROOT / "storage" / "runs" / f"user_{user_id}" / run_id
    for child in ("input", "registry", "output", "logs", "errors"):
        (run_dir / child).mkdir(parents=True, exist_ok=True)
    return run_dir


def _source_path(run_dir: Path, original_file_name: str | None) -> Path:
    suffix = Path(original_file_name or "").suffix or ".xlsx"
    return run_dir / "input" / f"source{suffix}"


def _object_addresses_path() -> Path:
    settings = load_settings()
    return settings.object_addresses_path or DEFAULT_OBJECT_ADDRESSES_PATH


def _court_orders_base_data_path() -> Path:
    settings = load_settings()
    return settings.court_orders_base_data_path or DEFAULT_COURT_ORDERS_BASE_DATA_PATH


def _court_orders_jurisdiction_path() -> Path:
    settings = load_settings()
    return (
        settings.court_orders_jurisdiction_path
        or DEFAULT_COURT_ORDERS_JURISDICTION_PATH
    )


def _validate_document(document) -> str | None:
    suffix = Path(document.file_name or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        return "Загрузите файл Excel в формате .xlsx или .xlsm."

    max_upload_mb = load_settings().max_upload_mb
    if document.file_size and document.file_size > max_upload_mb * 1024 * 1024:
        return f"Файл слишком большой. Максимальный размер: {max_upload_mb} МБ."

    return None


def _registry_error_count(registry_path: str) -> int:
    workbook = load_workbook(registry_path, read_only=True, data_only=True)
    if "Ошибки" not in workbook.sheetnames:
        return 0

    worksheet = workbook["Ошибки"]
    return max(worksheet.max_row - 1, 0)


def _registry_sheet_count(registry_path: str, sheet_name: str) -> int:
    workbook = load_workbook(registry_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return 0
    return max(workbook[sheet_name].max_row - 1, 0)


def _zip_docx_count(zip_path: str) -> int:
    with ZipFile(zip_path) as archive:
        return sum(name.lower().endswith(".docx") for name in archive.namelist())


def _install_object_addresses(upload_path: Path) -> Path:
    target_path = _object_addresses_path()
    target_path.parent.mkdir(parents=True, exist_ok=True)
    load_object_addresses(upload_path)
    upload_path.replace(target_path)
    return target_path


def _install_court_orders_base_data(
    upload_path: Path,
) -> tuple[WorkbookValidationReport, bool]:
    report, backup_path = install_validated_reference_file(
        upload_path,
        _court_orders_base_data_path(),
        validate_base_data,
    )
    return report, backup_path is not None


def _install_court_orders_jurisdiction(
    upload_path: Path,
) -> tuple[WorkbookValidationReport, bool]:
    report, backup_path = install_validated_reference_file(
        upload_path,
        _court_orders_jurisdiction_path(),
        validate_jurisdiction_data,
    )
    return report, backup_path is not None


def _format_reference_report(
    report: WorkbookValidationReport,
    backup_created: bool,
) -> str:
    lines = [
        f"Готово: {report.kind} обновлена.",
        f"Объектов: {report.objects_count}.",
    ]
    if report.warning_counts:
        lines.append("Предупреждения:")
        lines.extend(
            f"• {label}: {count}"
            for label, count in report.warning_counts.items()
        )
    else:
        lines.append("Предупреждений нет.")
    if backup_created:
        lines.append("Предыдущая версия сохранена в резервных копиях.")
    return "\n".join(lines)


async def _send_menu(message: Message) -> None:
    user_id = message.from_user.id if message.from_user else None
    await message.answer("Выберите действие:", reply_markup=_main_menu(user_id))


@router.message(CommandStart())
async def start(message: Message, state: FSMContext) -> None:
    user_id = message.from_user.id if message.from_user else None
    logger.info("Start from user_id=%s", user_id)
    if await _deny_if_needed(message):
        return
    await state.clear()
    await _send_menu(message)


@router.message(Command("dictionary"))
async def upload_dictionary(message: Message, state: FSMContext) -> None:
    if await _deny_if_needed(message) or await _deny_admin_message_if_needed(message):
        return
    await state.update_data(action="dictionary")
    await state.set_state(DocumentFlow.waiting_for_file)
    await message.answer(ACTION_PROMPTS["dictionary"])


@router.message(Command("courtdata"))
async def upload_court_data(message: Message, state: FSMContext) -> None:
    if await _deny_if_needed(message) or await _deny_admin_message_if_needed(message):
        return
    await state.update_data(action="court_data")
    await state.set_state(DocumentFlow.waiting_for_file)
    await message.answer(ACTION_PROMPTS["court_data"])


@router.message(Command("jurisdiction"))
async def upload_jurisdiction(message: Message, state: FSMContext) -> None:
    if await _deny_if_needed(message) or await _deny_admin_message_if_needed(message):
        return
    await state.update_data(action="jurisdiction")
    await state.set_state(DocumentFlow.waiting_for_file)
    await message.answer(ACTION_PROMPTS["jurisdiction"])


@router.callback_query(F.data == "menu:data_updates")
async def show_data_updates_menu(
    callback: CallbackQuery,
    state: FSMContext,
) -> None:
    if (
        await _deny_callback_if_needed(callback)
        or await _deny_admin_callback_if_needed(callback)
    ):
        return
    await state.clear()
    await callback.message.edit_text(
        "Обновление данных:",
        reply_markup=data_updates_menu_keyboard(),
    )
    await _answer_callback_safely(callback)


@router.callback_query(F.data == "menu:main")
async def show_main_menu(
    callback: CallbackQuery,
    state: FSMContext,
) -> None:
    if await _deny_callback_if_needed(callback):
        return
    await state.clear()
    await callback.message.edit_text(
        "Выберите действие:",
        reply_markup=_main_menu(callback.from_user.id if callback.from_user else None),
    )
    await _answer_callback_safely(callback)


@router.callback_query(F.data == "action:help")
async def show_help(callback: CallbackQuery, state: FSMContext) -> None:
    if await _deny_callback_if_needed(callback):
        return
    await state.clear()
    await callback.message.answer(
        HELP_TEXT,
        reply_markup=_main_menu(callback.from_user.id if callback.from_user else None),
    )
    await _answer_callback_safely(callback)


@router.callback_query(F.data.in_({"action:normalize", "action:claims", "action:dictionary", "action:court_data", "action:jurisdiction", "action:court_orders"}))
async def choose_action(callback: CallbackQuery, state: FSMContext) -> None:
    if await _deny_callback_if_needed(callback):
        return

    action = callback.data.split(":", 1)[1]
    if action in REFERENCE_ACTIONS and await _deny_admin_callback_if_needed(callback):
        return
    await state.update_data(action=action)
    await state.set_state(DocumentFlow.waiting_for_file)
    await callback.message.answer(ACTION_PROMPTS[action])
    await _answer_callback_safely(callback)


@router.message(DocumentFlow.waiting_for_file, F.document)
async def receive_document(message: Message, bot: Bot, state: FSMContext) -> None:
    if await _deny_if_needed(message):
        return

    data = await state.get_data()
    action = data.get("action")
    if action not in {"normalize", "claims", "dictionary", "court_data", "jurisdiction", "court_orders"}:
        await state.clear()
        await message.answer("Выберите действие заново.", reply_markup=main_menu_keyboard())
        return
    if action in REFERENCE_ACTIONS and await _deny_admin_message_if_needed(message):
        await state.clear()
        return

    document = message.document
    validation_error = _validate_document(document)
    if validation_error:
        await message.answer(validation_error, reply_markup=_main_menu(message.from_user.id))
        await state.clear()
        return

    run_dir = _create_run_dir(message.from_user.id)
    input_path = _source_path(run_dir, document.file_name)

    await message.answer("Файл получен. Обрабатываю...")
    await bot.download(document, destination=input_path)
    logger.info(
        "Audit action_started user_id=%s action=%s run_id=%s",
        message.from_user.id,
        action,
        run_dir.name,
    )

    try:
        if action == "dictionary":
            target_path = await asyncio.to_thread(_install_object_addresses, input_path)
            await message.answer(
                f"Готово: справочник адресов обновлен.\nПуть: {target_path}",
                reply_markup=_main_menu(message.from_user.id),
            )
        elif action == "court_data":
            report, backup_created = await asyncio.to_thread(
                _install_court_orders_base_data,
                input_path,
            )
            await message.answer(
                _format_reference_report(report, backup_created),
                reply_markup=_main_menu(message.from_user.id),
            )
        elif action == "jurisdiction":
            report, backup_created = await asyncio.to_thread(
                _install_court_orders_jurisdiction,
                input_path,
            )
            await message.answer(
                _format_reference_report(report, backup_created),
                reply_markup=_main_menu(message.from_user.id),
            )
        elif action == "normalize":
            result_path = await asyncio.to_thread(run_excel_to_registry, str(input_path), str(run_dir))
            error_count = await asyncio.to_thread(_registry_error_count, result_path)
            good_payers_count = await asyncio.to_thread(
                _registry_sheet_count,
                result_path,
                "Добросовестные плательщики",
            )
            caption = "Готово: registry.xlsx"
            if error_count:
                caption += f"\nЕсть строки на листе «Ошибки»: {error_count}."
            if good_payers_count:
                caption += (
                    "\nДобросовестных плательщиков: "
                    f"{good_payers_count}."
                )
            await message.answer_document(
                FSInputFile(result_path),
                caption=caption,
            )
        elif action == "claims":
            result_path = await asyncio.to_thread(run_registry_to_claims, str(input_path), str(run_dir))
            documents_count = await asyncio.to_thread(_zip_docx_count, result_path)
            await message.answer_document(
                FSInputFile(result_path),
                caption=f"Готово: claims.zip\nСоздано претензий: {documents_count}.",
            )
        elif action == "court_orders":
            result_path = await asyncio.to_thread(run_registry_to_court_orders, str(input_path), str(run_dir))
            documents_count = await asyncio.to_thread(_zip_docx_count, result_path)
            await message.answer_document(
                FSInputFile(result_path),
                caption=(
                    "Готово: court_orders.zip\n"
                    f"Создано заявлений: {documents_count}.\n"
                    "Проверьте errors.xlsx, если он включён в архив."
                ),
            )
    except Exception as exc:
        logger.error(
            "Audit action_failed user_id=%s action=%s run_id=%s error_type=%s",
            message.from_user.id,
            action,
            run_dir.name,
            type(exc).__name__,
        )
        error_text = (
            str(exc)
            if isinstance(exc, (ValueError, FileNotFoundError))
            else "Ошибка обработки файла. Проверьте формат и попробуйте еще раз."
        )
        await message.answer(
            error_text,
            reply_markup=_main_menu(message.from_user.id),
        )
        return
    finally:
        await state.clear()

    logger.info(
        "Audit action_completed user_id=%s action=%s run_id=%s",
        message.from_user.id,
        action,
        run_dir.name,
    )
    await _send_menu(message)


@router.message(DocumentFlow.waiting_for_file)
async def receive_non_document(message: Message) -> None:
    if await _deny_if_needed(message):
        return
    await message.answer("Пожалуйста, загрузите Excel-файл документом.")
