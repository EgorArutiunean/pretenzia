#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_debt_registry.py

Скрипт для преобразования иерархического отчёта 1С по начислениям/оплатам
в Excel-реестр для генерации претензий в формате шаблона.

Ожидаемая структура отчёта 1С:
- level 3: договор / лицевой счёт, например "12970001 от 15.09.2020"
- level 4: контрагент / должник
- level 5: периоды, например "Январь 2026"
- колонка E: Начислено
- колонка F: Оплачено
- колонка G: Долг на конец периода

Пример запуска:
python build_debt_registry.py "onv_report.xlsx" --object-address "Адрес объекта" --out "registry.xlsx"
"""

import argparse
import json
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from app.shared.object_data import load_object_data


MONTHS_RU = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def to_decimal(value):
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(
            str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".")
        )
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Не удалось распознать денежную сумму: {value!r}") from exc


def parse_account_number(contract_text):
    """
    Берём первый набор цифр в начале строки договора.
    Пример: "12970001 от 15.09.2020" -> "12970001"
    """
    if not contract_text:
        return ""
    match = re.match(r"^\s*(\d+)", str(contract_text))
    return match.group(1) if match else ""


def parse_contract_date(contract_text):
    """
    Достаём дату после "от".
    Пример: "12970001 от 15.09.2020" -> "15.09.2020"
    """
    if not contract_text:
        return ""
    match = re.search(r"(\d{2}\.\d{2}\.\d{4})", str(contract_text), flags=re.IGNORECASE)
    return match.group(1) if match else ""


def parse_period(period_text):
    """
    Пример: "Январь 2026" -> datetime(2026, 1, 1)
    """
    if not period_text:
        return None

    text = str(period_text).strip().lower()
    parts = text.split()
    if len(parts) < 2:
        return None

    month = MONTHS_RU.get(parts[0])
    try:
        year = int(parts[1])
    except ValueError:
        return None

    if not month:
        return None

    return datetime(year, month, 1)


def format_period_range(periods):
    """
    Пример:
    [2025-11-01, 2026-02-01] -> "01.11.2025 — 01.02.2026"
    """
    periods = [p for p in periods if p]
    if not periods:
        return ""

    start = min(periods)
    end = max(periods)
    return f"{start:%d.%m.%Y} — {end:%d.%m.%Y}"


def normalize_object_address(object_address, parking_place_number):
    object_address = (object_address or "").strip().rstrip(",")
    if parking_place_number:
        return f"{object_address}, машиноместо № {parking_place_number}"
    return object_address


def parse_account_parts(account_number):
    digits = re.sub(r"\D+", "", str(account_number or ""))
    if len(digits) != 8:
        return "", ""
    return digits[:4], digits[4:]


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def load_object_addresses(path):
    if not path:
        return {}

    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Справочник адресов объектов не найден: {source}")
    if not source.is_file():
        raise ValueError(f"Справочник адресов объектов должен быть файлом .xlsx или .json: {source}")

    if source.suffix.lower() == ".json":
        with source.open("r", encoding="utf-8") as file:
            data = json.load(file)
        if isinstance(data, dict):
            return {str(code).zfill(4): str(address).strip() for code, address in data.items() if str(address).strip()}
        if isinstance(data, list):
            result = {}
            for item in data:
                if not isinstance(item, dict):
                    continue
                code = item.get("object_code") or item.get("code") or item.get("Код объекта")
                address = item.get("object_address") or item.get("address") or item.get("Адрес объекта")
                if code and address:
                    result[str(code).zfill(4)] = str(address).strip()
            return result
        raise ValueError("JSON-справочник должен быть объектом {object_code: object_address} или списком объектов.")

    if source.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(source, data_only=True)
        ws = wb.active
        header_row = 1
        code_col = None
        address_col = None

        for row in range(1, min(ws.max_row, 20) + 1):
            headers = {normalize_header(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1)}
            code_col = (
                headers.get("object_code")
                or headers.get("код объекта")
                or headers.get("код")
                or headers.get("№")
                or headers.get("номер")
            )
            address_col = headers.get("object_address") or headers.get("адрес объекта") or headers.get("адрес")
            if code_col and address_col:
                header_row = row
                break

        if not code_col or not address_col:
            code_col, address_col = 1, 2

        result = {}
        for row in range(header_row + 1, ws.max_row + 1):
            code = ws.cell(row, code_col).value
            address = ws.cell(row, address_col).value
            if code and address:
                result[str(code).strip().zfill(4)] = str(address).strip()
        return result

    raise ValueError("Справочник адресов должен быть в формате .xlsx, .xlsm или .json.")


def is_contract_text(value):
    if not value:
        return False
    text = str(value).strip()
    return bool(re.match(r"^\d+", text) and re.search(r"\d{2}\.\d{2}\.\d{4}", text))


def detect_amount_columns(ws):
    charged_col = 5
    debt_col = 7

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            text = value.strip().lower()
            if text == "начислено":
                charged_col = cell.column
            elif "долг" in text and "конец" in text:
                debt_col = cell.column

    return charged_col, debt_col


def build_registry(
    input_path,
    output_path,
    object_address=None,
    sheet_name=None,
    min_debt=0,
    group_by="contract",
    object_addresses_path=None,
    base_data_path=None,
):
    wb = load_workbook(input_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    charged_col, debt_col = detect_amount_columns(ws)
    object_address_by_code = load_object_addresses(object_addresses_path)
    object_data_by_code = load_object_data(base_data_path) if base_data_path else {}

    rows = []
    errors = []
    good_payers = []
    object_addresses = set()

    current_object_address = ""
    current_org = ""
    current_contract_text = ""
    current_account = ""
    current_contract_date = ""
    current_contract_debt = Decimal("0")
    current_debtor = ""
    current_periods_positive_balance = []
    current_periods_with_charge = []
    current_source_row = 0
    current_amount_error = ""

    def flush_current():
        nonlocal current_contract_text, current_account, current_contract_date
        nonlocal current_contract_debt, current_debtor
        nonlocal current_periods_positive_balance, current_periods_with_charge
        nonlocal current_source_row, current_amount_error

        if not current_contract_text or not current_debtor:
            return

        if current_amount_error:
            errors.append({
                "source_row": current_source_row,
                "account_number": current_account,
                "debtor_name": current_debtor,
                "company": current_org,
                "contract_text": current_contract_text,
                "reason": current_amount_error,
            })
            return

        debt = current_contract_debt
        account = current_account
        object_code, parking_place_number = parse_account_parts(account)
        if not object_code:
            errors.append({
                "source_row": current_source_row,
                "account_number": account,
                "debtor_name": current_debtor,
                "company": current_org,
                "contract_text": current_contract_text,
                "reason": "Лицевой счёт должен состоять ровно из 8 цифр.",
            })
            return

        object_data = object_data_by_code.get(object_code)
        if object_data_by_code and object_data is None:
            errors.append({
                "source_row": current_source_row,
                "account_number": account,
                "debtor_name": current_debtor,
                "company": current_org,
                "contract_text": current_contract_text,
                "reason": f"Код объекта {object_code} отсутствует в основной БЗ.",
            })
            return
        if object_data is not None and not object_data.company:
            errors.append({
                "source_row": current_source_row,
                "account_number": account,
                "debtor_name": current_debtor,
                "company": "",
                "contract_text": current_contract_text,
                "reason": f"Для кода объекта {object_code} не указана компания.",
            })
            return
        if object_address_by_code:
            row_object_address = object_address_by_code.get(object_code)
            if not row_object_address:
                errors.append({
                    "source_row": current_source_row,
                    "account_number": account,
                    "debtor_name": current_debtor,
                    "company": object_data.company if object_data else current_org,
                    "contract_text": current_contract_text,
                    "reason": f"Не найден адрес объекта для кода {object_code or '----'}",
                })
                return
        else:
            row_object_address = (
                object_data.object_address
                if object_data is not None
                else object_address or current_object_address
            )

        organization = object_data.company if object_data is not None else current_org
        company_inn = object_data.inn if object_data is not None else ""
        company_code = object_data.company_code if object_data is not None else ""

        # Основное правило периода:
        # берём месяцы, где долг на конец периода положительный.
        # Если таких нет, запасной вариант — месяцы с начислением.
        period_source = current_periods_positive_balance or current_periods_with_charge
        if row_object_address:
            object_addresses.add(row_object_address)

        normalized_row = {
            "source_row": current_source_row,
            "organization": organization,
            "company_inn": company_inn,
            "company_code": company_code,
            "account_number": account,
            "contract_text": current_contract_text,
            "contract_date": current_contract_date,
            "debtor_name": current_debtor,
            "object_address": row_object_address,
            "parking_place_number": parking_place_number,
            "full_object_address": normalize_object_address(row_object_address, parking_place_number),
            "debt_period": format_period_range(period_source),
            "debt_amount": debt,
        }
        if debt <= Decimal("0"):
            good_payers.append(normalized_row)
            return
        if debt <= Decimal(str(min_debt)):
            return
        rows.append(normalized_row)

    for r in range(1, ws.max_row + 1):
        value = ws.cell(r, 1).value
        if value is None:
            continue

        value = str(value).strip()
        level = ws.row_dimensions[r].outlineLevel
        period_date = parse_period(value)

        if period_date and current_contract_text:
            try:
                debt_end = to_decimal(ws.cell(r, debt_col).value)
                charged = to_decimal(ws.cell(r, charged_col).value)
            except ValueError as exc:
                current_amount_error = f"Строка {r}: {exc}"
                continue
            if charged != 0:
                current_periods_with_charge.append(period_date)
            if debt_end > 0:
                current_periods_positive_balance.append(period_date)
            continue

        if is_contract_text(value):
            flush_current()

            current_contract_text = value
            current_account = parse_account_number(value)
            current_contract_date = parse_contract_date(value)
            current_source_row = r
            current_amount_error = ""
            try:
                current_contract_debt = to_decimal(ws.cell(r, debt_col).value)
            except ValueError as exc:
                current_contract_debt = Decimal("0")
                current_amount_error = f"Строка {r}: {exc}"
            current_debtor = ""
            current_periods_positive_balance = []
            current_periods_with_charge = []
            continue

        if current_contract_text and not current_debtor and level > 0:
            current_debtor = value
            if current_contract_debt == 0 and not current_amount_error:
                try:
                    debt_end = to_decimal(ws.cell(r, debt_col).value)
                except ValueError as exc:
                    current_amount_error = f"Строка {r}: {exc}"
                else:
                    if debt_end != 0:
                        current_contract_debt = debt_end
            continue

        if level == 1 and not current_contract_text:
            current_object_address = value
            if not object_address_by_code:
                object_addresses.add(value)
            continue

        if level in (2, 3) and not current_contract_text:
            current_org = value
            continue

    flush_current()

    unique_rows = []
    seen_rows = set()
    for row in rows:
        key = tuple(
            row[field]
            for field in (
                "organization",
                "company_inn",
                "company_code",
                "account_number",
                "contract_text",
                "debtor_name",
                "full_object_address",
                "debt_period",
                "debt_amount",
            )
        )
        if key in seen_rows:
            continue
        seen_rows.add(key)
        unique_rows.append(row)
    rows = unique_rows

    if group_by == "account":
        grouped = {}
        for row in rows:
            key = row["account_number"]
            if key not in grouped:
                grouped[key] = row.copy()
                grouped[key]["debt_amount"] = Decimal("0")
                grouped[key]["_period_dates"] = []
                grouped[key]["contract_text"] = []
            grouped[key]["debt_amount"] += row["debt_amount"]
            grouped[key]["contract_text"].append(row["contract_text"])
            # восстановим даты из debt_period
            for dt_str in re.findall(r"\d{2}\.\d{2}\.\d{4}", row["debt_period"]):
                try:
                    grouped[key]["_period_dates"].append(datetime.strptime(dt_str, "%d.%m.%Y"))
                except ValueError:
                    pass

        rows = []
        for row in grouped.values():
            row["contract_text"] = "; ".join(row["contract_text"])
            row["debt_period"] = format_period_range(row.pop("_period_dates", []))
            rows.append(row)

    rows.sort(key=lambda x: (x["account_number"], x["debtor_name"]))

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Реестр"

    # Формат выходного файла строго как в шаблоне реестра:
    # Лицевой счет | ФИО | Адрес | Период задолженности | Сумма долга
    headers = [
        "Лицевой счет",
        "ФИО",
        "Адрес",
        "Период задолженности",
        "Сумма долга",
        "Компания",
        "ИНН компании",
        "Код компании",
    ]

    out_ws.append(headers)

    for row in rows:
        out_ws.append([
            row["account_number"],
            row["debtor_name"],
            row["full_object_address"],
            row["debt_period"].replace(" — ", "-"),
            row["debt_amount"],
            row["organization"],
            row["company_inn"],
            row["company_code"],
        ])

    # Оформление
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in out_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in out_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, width in {
        1: 16,
        2: 34,
        3: 55,
        4: 24,
        5: 16,
        6: 34,
        7: 16,
        8: 16,
    }.items():
        out_ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Лицевой счёт храним как текст, чтобы Excel не менял формат.
    for cell in out_ws["A"][1:]:
        cell.number_format = '@'

    for cell in out_ws["E"][1:]:
        cell.number_format = '#,##0.00'

    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = out_ws.dimensions

    # Сводка
    summary = out_wb.create_sheet("Сводка")
    summary.append(["Показатель", "Значение"])
    summary.append(["Входной файл", input_path])
    summary.append(["Адрес объекта", object_address or ", ".join(sorted(object_addresses))])
    summary.append(["Количество претензий", len(rows)])
    summary.append(["Количество ошибок", len(errors)])
    summary.append(["Итоговая сумма долга", sum(r["debt_amount"] for r in rows)])
    summary.append(["Группировка", group_by])
    if object_addresses_path:
        summary.append(["Справочник адресов", object_addresses_path])
    if base_data_path:
        summary.append(["Основная БЗ", base_data_path])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 70
    for cell in summary["B"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

    errors_ws = out_wb.create_sheet("Ошибки")
    errors_ws.append(
        ["Строка ОНВ", "Лицевой счет", "ФИО", "Компания", "Договор", "Причина"]
    )
    for error in errors:
        errors_ws.append([
            error.get("source_row"),
            error["account_number"],
            error["debtor_name"],
            error.get("company", ""),
            error["contract_text"],
            error["reason"],
        ])

    for cell in errors_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in errors_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, width in {1: 12, 2: 16, 3: 34, 4: 34, 5: 32, 6: 55}.items():
        errors_ws.column_dimensions[get_column_letter(col_idx)].width = width

    good_ws = out_wb.create_sheet("Добросовестные плательщики")
    good_ws.append(
        [
            "Строка ОНВ",
            "Лицевой счет",
            "ФИО",
            "Компания",
            "Сумма долга",
            "Причина",
        ]
    )
    for row in good_payers:
        good_ws.append(
            [
                row["source_row"],
                row["account_number"],
                row["debtor_name"],
                row["organization"],
                row["debt_amount"],
                "Задолженность меньше или равна нулю.",
            ]
        )
    for cell in good_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row in good_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, width in {1: 12, 2: 16, 3: 34, 4: 34, 5: 16, 6: 42}.items():
        good_ws.column_dimensions[get_column_letter(col_idx)].width = width
    for cell in good_ws["E"][1:]:
        cell.number_format = '#,##0.00'

    out_wb.save(output_path)
    return rows


def main():
    parser = argparse.ArgumentParser(description="Преобразует отчёт 1С в реестр для претензий.")
    parser.add_argument("input", help="Путь к входному Excel-файлу из 1С")
    parser.add_argument("--out", default="debt_registry.xlsx", help="Путь к выходному Excel-файлу")
    parser.add_argument("--object-addresses", default=None, help="Справочник адресов объектов .xlsx или .json: object_code -> object_address")
    parser.add_argument("--base-data", default=None, help="Основная БЗ объектов и компаний .xlsx")
    parser.add_argument("--object-address", default=None, help="Адрес объекта / паркинга. Если не указан, берётся из отчёта.")
    parser.add_argument("--sheet", default=None, help="Имя листа. Если не указано, берётся первый лист.")
    parser.add_argument("--min-debt", type=float, default=0, help="Минимальная сумма долга для включения в реестр")
    parser.add_argument(
        "--group-by",
        choices=["contract", "account"],
        default="contract",
        help="contract = отдельная строка на договор; account = объединять по лицевому счёту",
    )

    args = parser.parse_args()

    rows = build_registry(
        input_path=args.input,
        output_path=args.out,
        object_address=args.object_address,
        sheet_name=args.sheet,
        min_debt=args.min_debt,
        group_by=args.group_by,
        object_addresses_path=args.object_addresses,
        base_data_path=args.base_data,
    )

    print(f"Готово: {args.out}")
    print(f"Строк в реестре: {len(rows)}")
    print(f"Итоговый долг: {sum(row['debt_amount'] for row in rows):,.2f}")


if __name__ == "__main__":
    main()
